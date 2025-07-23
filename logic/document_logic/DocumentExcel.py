import json
from openpyxl import Workbook
from logic.document_logic.document_text_logic import Documents
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

class DocumentExcel(Documents):
  """
  Initialize  the class with the name, and create a path with that.
  """
  def __init__(self, name):
    self.path = f"data_json/{name}.json"
    self.data = []
    super().__init__(name=self.path, type="Excel")

  def load_data(self):
    """
    Search a .json file, and try to extract it data to work with it.
    """
    try:
      with open(self.path, encoding="utf-8") as f:
        self.data = json.load(f)
        if isinstance(self.data, list):
          print("Something still wrong, why is this a list")
        else:
          print("JSON loaded, this are the tags:", list(self.data.keys()))
    except FileNotFoundError:
        print(f"Load of {self.path}, failed")

  def normalize_data(self):
    """
    Compress the data and prepare it for the excel document
    """
    new_data = []

    # Tags like: name_product, price_product, url_product, etc.
    fields = list(self.data.keys())
    num_items = len(self.data[fields[0]])

    for i in range(num_items):
        item = {"id": str(i + 1)}
        for field in fields:
            try:
                field_data = self.data[field][i]      
                outer_key = next(iter(field_data))   
                inner_dict = field_data[outer_key]   
                inner_key = next(iter(inner_dict))  
                value = inner_dict[inner_key]          
                item[field] = value
            except Exception:
                item[field] = "<<NO DATA>>"

        new_data.append(item)

    self.data = new_data

  def create_document(self, title: "str"):
    """
    Creates a .xlsx document with the fixed data coming from normalize_data()
    Creates a styilized excel, readable with all the tags
    """
    if not self.data:
        raise FileNotFoundError("No data found")

    excel = Workbook()
    page = excel.active
    page.title = title

    # Headers
    heads = list(self.data[0].keys())
    page.append(heads)

    # Headers style
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, _ in enumerate(heads, start=1):
        cell = page.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = center_alignment

    # Row of data (All info of each item)
    for item in self.data:
        row = []
        for key in heads:
            row.append(item.get(key, ""))
        page.append(row)

    # Cells Style
    for col_num, col in enumerate(page.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in col:
            cell.alignment = center_alignment
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        page.column_dimensions[column_letter].width = adjusted_width

    # Guardar archivo
    excel.save(f"{title}.xlsx")
